import openpyxl
import sys

wb = openpyxl.load_workbook('Endpoint_desc.xlsx')
ws = wb.active

# Print headers and existing rows to understand structure
print("=== HEADERS ===")
headers = [cell.value for cell in ws[1]]
for i, h in enumerate(headers):
    print(f"  Col {i}: {repr(h)}")

print(f"\n=== EXISTING ROWS ({ws.max_row - 1} rows) ===")
for row in ws.iter_rows(min_row=2, values_only=True):
    print(row)
